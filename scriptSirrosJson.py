from openpyxl import load_workbook
import json
import re

# ---------- carregar produtos.json ----------
with open("produtos_referencias.json", "r", encoding="utf-8") as f:
    produtos_base = json.load(f)

produtos_index = {p["Referencia"]: p for p in produtos_base}

# ---------- carregar excel ----------
wb = load_workbook("itens.xlsx")
ws = wb.active

# localizar cabeçalho
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row and "Código" in row:
        header_row = i
        headers = list(row)
        break

idx = {h: i for i, h in enumerate(headers)}

resultado = []
nao_encontrados = set()

for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    if not row or not row[idx["Código"]]:
        continue

    referencia = row[idx["Código"]]
    descricao = row[idx["Descrição"]]

    preco_raw = row[idx["Venda"]]

    if not preco_raw or preco_raw == "Venda":
        continue

    try:
        if isinstance(preco_raw, str):
            preco = float(
                re.sub(r"[R$\s]", "", preco_raw)
            )
        else:
            preco = float(preco_raw)
    except ValueError:
        continue

    base = produtos_index.get(referencia)

    if not base:
        nao_encontrados.add(referencia)

    resultado.append({
        "Referencia": referencia,
        "Descricao": descricao,
        "Categoria": base.get("Categoria") if base else None,
        "Classificacao": base.get("Classificacao") if base else None,
        "Tipo_Produto": "Tem Componentes, Componente",
        "Preco": preco
    })

# ---------- salvar ----------
with open("produtos.json", "w", encoding="utf-8") as f:
    json.dump(resultado, f, ensure_ascii=False, indent=2)

with open("referencias_nao_encontradas.txt", "w", encoding="utf-8") as f:
    for ref in sorted(nao_encontrados):
        f.write(ref + "\n")
